'''

Modules that must be installed on the system running this script:

    openpyxl (3.1.5)
    matplotlib (3.7.2)
    numpy (1.26.4)
    pygame (2.6.1)
    scipy (1.15.3)


To ensure this script runs correctly, run this command in the Windows terminal:

    python -m pip install --upgrade openpyxl==3.1.5 matplotlib==3.7.2 numpy==1.26.4 pygame==2.6.1 scipy==1.15.3

'''

import os
os.system('cls')
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = '1'

import warnings
warnings.filterwarnings('ignore', message='pkg_resources is deprecated as an API')

import time
from datetime import datetime, timedelta, date, time as dtime
from io import BytesIO
from itertools import cycle
import json
import csv

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from openpyxl import load_workbook
import numpy as np
import pygame as pg
from scipy.signal import butter, filtfilt
from scipy.spatial.distance import pdist, squareform


# -----------------------------
# Data Processing
# -----------------------------
def _filter_trials(easy_rates, outcomes, miss_threshold=4):
    K1, K2 = easy_rates
    N1, N2 = 4 * K1, 4 * K2

    min_idx = N1 + N2 - 1

    if len(outcomes) == 0:
        return 0
    if min_idx >= len(outcomes):
        return len(outcomes)

    consecutive_misses = 0
    stop_idx = len(outcomes) - 1

    for i, o in enumerate(outcomes[min_idx:], start=min_idx):
        if o == 'miss':
            consecutive_misses += 1
        else:
            consecutive_misses = 0
        
        if consecutive_misses >= miss_threshold:
            stop_idx = i - miss_threshold
            break
    
    return stop_idx + 1


def _interp(times, values):
    target_dt = 0.001
    min_duration = 30.0

    if times is None or values is None:
        return [], np.array([], dtype=float)
    
    if len(times) == 0 or len(values) == 0:
        return [], np.array([], dtype=float)
    
    if len(times) != len(values):
        raise ValueError('times and values must have the same length')
    
    values = np.asarray(values, dtype=float)

    n = len(times)
    if n == 1:
        new_times = [times[0]]
        new_values = [float(values[0])]

        start_t = times[0]
        stop_t = start_t + timedelta(seconds=min_duration)

        last_t = new_times[-1]
        last_v = new_values[-1]

        while (stop_t - last_t).total_seconds() > 0:
            last_t = last_t + timedelta(seconds=target_dt)

            new_times.append(last_t)
            new_values.append(last_v)
        
        return new_times, np.asarray(new_values, dtype=float)
    
    init_start = times[0]
    init_stop = times[-1]

    new_times = [times[0]]
    new_values = [float(values[0])]

    step_td = timedelta(seconds=target_dt)

    for i in range(1, n):
        t_curr = times[i]
        v_prev = float(values[i-1])
        v_curr = float(values[i])

        last_t = new_times[-1]
        dt = (t_curr - last_t).total_seconds()

        if dt <= 0:
            new_times.append(t_curr)
            new_values.append(v_curr)
            continue

        while dt > target_dt:
            last_t = last_t + step_td

            new_times.append(last_t)
            new_values.append(v_prev)

            dt = (t_curr - last_t).total_seconds()
        
        new_times.append(t_curr)
        new_values.append(v_curr)
    
    init_duration = (init_stop - init_start).total_seconds()
    if init_duration < min_duration:
        target_stop = init_start + timedelta(seconds=min_duration)
        last_t = new_times[-1]
        last_v = new_values[-1]

        while (target_stop - last_t).total_seconds() > 0:
            last_t = last_t + step_td

            new_times.append(last_t)
            new_values.append(last_v)
    
    return new_times, np.asarray(new_values, dtype=float)


def _fft(times, values):
    if times is None or values is None:
        print('_fft: times or values is None')
        return

    if len(times) == 0 or len(values) == 0:
        print('_fft: empty times or values')
        return

    if len(times) != len(values):
        raise ValueError('times and values must have the same length')

    pairs = [(t, v) for (t, v) in zip(times, values) if v is not None]
    if len(pairs) < 2:
        print('_fft: not enough valid samples')
        return

    times = [t for (t, _) in pairs]
    values = [v for (_, v) in pairs]

    times_i, vals_i = _interp(times, values)
    if len(times_i) < 2:
        print('_fft: not enough samples after interpolation')
        return

    times_s = np.array([(t - times_i[0]).total_seconds() for t in times_i],
                       dtype=float)
    dt_arr = np.diff(times_s)
    dt_arr = dt_arr[dt_arr > 0]
    if dt_arr.size == 0:
        print('_fft: invalid time axis')
        return

    dt = float(np.median(dt_arr))
    if dt <= 0.0 or not np.isfinite(dt):
        print('_fft: non-positive dt')
        return

    x = np.asarray(vals_i, dtype=float)
    if x.size < 2:
        print('_fft: not enough samples for FFT')
        return

    x = x - np.nanmean(x)

    N = x.size
    Y = np.fft.rfft(x)
    freqs = np.fft.rfftfreq(N, d=dt)

    mag = np.abs(Y)
    
    mag_min = float(np.min(mag))
    mag_max = float(np.max(mag))
    if mag_max > mag_min:
        mag_norm = (mag - mag_min) / (mag_max - mag_min)
    else:
        mag_norm = np.zeros_like(mag, dtype=float)

    for f, m in zip(freqs, mag_norm):
        if f > 0.1 and m >= 0.005:
            print(f'{f:.2f}')
            break

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(freqs, mag_norm, linewidth=1.0)

    ax.set_xlabel('Frequency (Hz)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Magnitude', fontsize=12, fontweight='bold')
    ax.set_title('One-sided FFT Magnitude', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3)

    fig.tight_layout()
    show_figure(fig)


def _lpf(x, fs):
    fc = np.asarray([0.1, 5.0], dtype=float)
    n = 1
    ftype = 'bandpass'

    x = np.asarray(x, dtype=float)

    if fs is None or fs <= 0 or not np.isfinite(fs):
        return x
    
    fN = 0.5 * fs
    if fN <= 0:
        return x
    
    wn = fc / fN
    if np.any(wn <= 0) or np.any(wn >= 1):
        return x
    
    try:
        b, a = butter(n, wn,
                      btype=ftype,
                      analog=False,
                      output='ba'
                      )
        return filtfilt(b, a, x)
    except Exception:
        return x


# -----------------------------
# Data Analysis
# -----------------------------
def _get_pct(*args, **kwargs):
    trials = kwargs.get('trials', args[0] if len(args) > 0 else None)
    outcomes = kwargs.get('outcomes', args[1] if len(args) > 1 else None)

    if not trials or not outcomes or len(trials) != len(outcomes):
        return []
    
    x = sum(1.0 if o == 'hit' else 0.0 for o in outcomes)
    N = len(outcomes)

    values = 100.0 * (x / N)
    
    return values


def _get_latency(*args, **kwargs):
    trials = kwargs.get('trials', args[0] if len(args) > 0 else None)
    outcomes = kwargs.get('outcomes', args[1] if len(args) > 1 else None)
    mode = kwargs.get('mode', 'hit')

    if not trials:
        return []
    
    if outcomes is None:
        outcomes = []
        for tr in trials:
            evt = tr.get('evt', [])
            outcomes.append(evt[-1][1] if evt else None)
    
    if len(outcomes) != len(trials):
        raise ValueError('trials and outcomes must have the same length')
    
    latencies = []
    for tr, outcome in zip(trials, outcomes):
        evt = tr['evt']
        if not evt:
            continue

        if mode in {'hit', 'miss'} and outcome != mode:
            continue

        t_start = evt[0][0]
        t_stop = None
        
        for t, e in evt:
            if e in {'hit', 'miss'}:
                t_stop = t
        
        if t_stop is None:
            continue

        dt = (t_stop - t_start).total_seconds()

        if dt > 30.0:
            dt = 30.0
        if dt >= 0.0:
            latencies.append(dt)
    
    return sorted(latencies)


def _get_displacement(*args, **kwargs):
    trials = kwargs.get('trials', args[0] if len(args) > 0 else None)
    outcomes = kwargs.get('outcomes', args[1] if len(args) > 1 else None)
    mode = kwargs.get('mode', 'miss')

    if not trials:
        return []
    
    if outcomes is None:
        outcomes = []
        for tr in trials:
            evt = tr.get('evt', [])
            outcomes.append(evt[-1][1] if evt else None)
    
    if len(outcomes) != len(trials):
        raise ValueError('trials and outcomes must have the same length')
    
    displacements = []

    for tr, outcome in zip(trials, outcomes):
        evt = tr.get('evt', [])
        enc = tr.get('enc', [])

        if not evt:
            continue

        if mode in {'hit', 'miss'} and outcome != mode:
            continue

        if enc:
            enc = [(t, v) for (t, v) in enc if v is not None]

        if not enc or len(enc) < 2:
            displacements.append(0.0)
            continue

        times = [t for (t, _) in enc]
        values = np.array([v for (_, v) in enc], dtype=float)

        times, values = _interp(times, values)

        times_s = np.array([(t - times[0]).total_seconds() for t in times], dtype=float)

        unique_times, unique_idx = np.unique(times_s, return_index=True)
        if unique_times.size < 2:
            displacements.append(0.0)
            continue

        vals_smooth = values[unique_idx].astype(float)

        valid_mask = np.isfinite(vals_smooth)
        if valid_mask.sum() >= 3:
            t_valid = unique_times[valid_mask]
            v_valid = vals_smooth[valid_mask]

            dt = np.min(np.diff(t_valid)) if np.min(np.diff(t_valid)) > 0 else np.median(np.diff(t_valid))
            if dt > 0:
                fs = 1.0 / dt

                v_filt = _lpf(v_valid, fs)

                vals_smooth = vals_smooth.copy()
                vals_smooth[valid_mask] = v_filt
        
        diffs = np.diff(vals_smooth)
        diffs = diffs[np.isfinite(diffs)]
        if diffs.size == 0:
            total_disp = 0.0
        else:
            total_disp = float(np.sum(np.abs(diffs)))
        
        displacements.append(total_disp)
    
    return sorted(displacements)


def _get_velocity(*args, **kwargs):  # [0, 600] rpm
    trials = kwargs.get('trials', args[0] if len(args) > 0 else None)
    outcomes = kwargs.get('outcomes', args[1] if len(args) > 1 else None)
    mode = kwargs.get('mode', 'all')
    method = kwargs.get('method', 'average')

    if not trials:
        return []
    
    if outcomes is None:
        outcomes = []
        for tr in trials:
            evt = tr.get('evt', [])
            outcomes.append(evt[-1][1] if evt else None)

    if len(outcomes) != len(trials):
        raise ValueError('trials and outcomes must have the same length')
    
    omegas = []

    for tr, outcome in zip(trials, outcomes):
        evt = tr.get('evt', [])
        enc = tr.get('enc', [])

        if not evt:
            continue

        if mode in {'hit', 'miss'} and outcome != mode:
            continue

        if not enc or len(enc) < 2:
            omegas.append(0.0)
            continue

        enc_valid = [(t, v) for (t, v) in enc if v is not None]
        if len(enc_valid) < 2:
            omegas.append(0.0)
            continue
        else:
            enc = enc_valid

        times = [t for (t, _) in enc]
        values = [v for (_, v) in enc]

        times_s = np.array([(t - times[0]).total_seconds() for t in times], dtype=float)
        vals = np.array(values, dtype=float)

        unique_times, unique_idx = np.unique(times_s, return_index=True)
        if unique_times.size < 2:
            omegas.append(0.0)
            continue

        unique_values = vals[unique_idx]

        vals_smooth = unique_values.astype(float)
        valid_mask = np.isfinite(vals_smooth)

        if valid_mask.sum() >= 3:
            t_valid = unique_times[valid_mask]
            v_valid = vals_smooth[valid_mask]

            dt = np.min(np.diff(t_valid)) if np.min(np.diff(t_valid)) > 0 else np.median(np.diff(t_valid))
            if dt > 0:
                fs = 1.0 / dt

                v_filt = _lpf(v_valid, fs)

                vals_smooth = vals_smooth.copy()
                vals_smooth[valid_mask] = v_filt

        vel = np.gradient(vals_smooth, unique_times)
        inst_vel = np.abs(vel) / 6.0
        
        if inst_vel.size:
            if method == 'average':
                o = float(np.mean(inst_vel))
            elif method == 'peak':
                o = float(np.max(inst_vel))
            elif method == 'percentile':
                o = float(np.percentile(inst_vel, 95))
            else:
                raise ValueError
        else:
            o = 0.0

        omegas.append(o)
    
    return sorted(omegas)


def _get_acceleration(*args, **kwargs):  # [0, 4e6] rev/min^2
    trials = kwargs.get('trials', args[0] if len(args) > 0 else None)
    outcomes = kwargs.get('outcomes', args[1] if len(args) > 1 else None)
    mode = kwargs.get('mode', 'hit')
    method = kwargs.get('method', 'percentile')

    if not trials:
        return []

    if outcomes is None:
        outcomes = []
        for tr in trials:
            evt = tr.get('evt', [])
            outcomes.append(evt[-1][1] if evt else None)
    
    if len(outcomes) != len(trials):
        raise ValueError('trials and outcomes must have the same length')
    
    alphas = []

    for tr, outcome in zip(trials, outcomes):
        evt = tr.get('evt', [])
        enc = tr.get('enc', [])

        if not evt:
            continue

        if mode in {'hit', 'miss'} and outcome != mode:
            continue

        if not enc or len(enc) < 3:
            alphas.append(0.0)
            continue
        
        enc_valid = [(t, v) for (t, v) in enc if v is not None]
        if len(enc_valid) < 3:
            alphas.append(0.0)
            continue
        else:
            enc = enc_valid

        times = [t for (t, _) in enc]
        values = [v for (_, v) in enc]

        times_s = np.array([(t - times[0]).total_seconds() for t in times], dtype=float)
        vals = np.array(values, dtype=float)

        unique_times, unique_idx = np.unique(times_s, return_index=True)
        if unique_times.size < 3:
            alphas.append(0.0)
            continue
        else:
            unique_values = vals[unique_idx]

        vals_smooth = unique_values.astype(float)
        valid_mask = np.isfinite(vals_smooth)

        if valid_mask.sum() >= 3:
            t_valid = unique_times[valid_mask]
            v_valid = vals_smooth[valid_mask]

            dt = np.min(np.diff(t_valid)) if np.min(np.diff(t_valid)) > 0 else np.median(np.diff(t_valid))
            if dt > 0:
                fs = 1.0 / dt

                v_filt = _lpf(v_valid, fs)

                vals_smooth = vals_smooth.copy()
                vals_smooth[valid_mask] = v_filt

        vel = np.gradient(vals_smooth, unique_times)
        acc = np.gradient(vel, unique_times)

        inst_alpha = np.abs(acc)

        if inst_alpha.size:
            if method == 'average':
                a = float(np.mean(inst_alpha))
            elif method == 'peak':
                a = float(np.max(inst_alpha))
            elif method == 'percentile':
                a = float(np.percentile(inst_alpha, 95))
            else:
                raise ValueError
        else:
            a = 0.0
        
        alphas.append(a * 10.0)  # revolutions / min^2

    return sorted(alphas)


def _get_duration(*args, **kwargs):
    trials = kwargs.get('trials', args[0] if len(args) > 0 else None)

    if not trials:
        return []
    
    elapsed = trials[0].get('elapsed')
    if elapsed is None:
        return []
    
    return elapsed


def _get_licks(*args, **kwargs):
    trials = kwargs.get('trials', args[0] if len(args) > 0 else None)
    outcomes = kwargs.get('outcomes', args[1] if len(args) > 1 else None)
    mode = kwargs.get('mode', 'all')
    method = kwargs.get('method', 'average')

    if not trials:
        return []
    
    if outcomes is None:
        outcomes = []
        for tr in trials:
            evt = tr.get('evt', [])
            out = None
            
            for _, e in evt:
                if e in {'hit', 'miss'}:
                    out = e

            if out is None and evt:
                out = evt[-1][1]
            
            outcomes.append(out)
    
    if len(outcomes) != len(trials):
        raise ValueError('trials and outcomes must have the same length')
    
    lick_dt = []

    for tr, outcome in zip(trials, outcomes):
        evt = tr.get('evt', [])
        if not evt:
            continue

        if mode in {'hit', 'miss'} and outcome != mode:
            continue

        t_outcome = None
        for t, e in evt:
            if e in {'hit', 'miss'}:
                t_outcome = t
        
        if t_outcome is None:
            continue

        deltas = []
        
        for t, e in evt:
            if e == 'lick':
                dt = (t - t_outcome).total_seconds()
                deltas.append(dt)
        
        if not deltas:
            continue

        deltas = np.asarray(deltas, dtype=float)

        if method == 'average':
            val = float(np.mean(deltas))
        elif method == 'median':
            val = float(np.median(deltas))
        
        lick_dt.append(val)
    
    return sorted(lick_dt)


# -----------------------------
# User Interaction
# -----------------------------
def _choose(prompt, options):
    while True:
        print(prompt)

        for i, opt in enumerate(options, start=1):
            print(f'\t[{i}] {opt}')
        
        ans = input('\nSelection: ').strip()

        try:
            idx = int(ans)
            if 1 <= idx <= len(options):
                print('\n')
                return options[idx-1]
        except Exception:
            pass
            
        print('Invalid selection, try again')


def _choose_multi(prompt, options):
    while True:
        print(prompt)

        for i, opt in enumerate(options, start=1):
            print(f'\t[{i}] {opt}')
        
        ans = input('\nSelect one or more: ').strip()
        if not ans:
            print('Must select at least one option\n')
            continue

        tokens = [token.strip() for token in ans.split(',') if token.strip()]
        if not tokens:
            print('Invalid entry, try again\n')
            continue

        groups = []
        
        if len(tokens) == 1 and '-' in tokens[0]:
            token = tokens[0]
            
            try:
                a, b = token.split('-', 1)
                a, b = map(int, (a, b))

                idx = list(range(min(a, b), max(a, b) + 1))
                idx = [i for i in idx if 1 <= i <= len(options)]
                if not idx:
                    print('No valid selections, try again\n')
                    continue

                for i in idx:
                    groups.append([options[i-1]])
            except Exception:
                print('Invalid entry, try again\n')
                continue
        else:
            try:
                for token in tokens:
                    if '-' in token:
                        a, b = token.split('-', 1)
                        a, b = map(int, (a, b))

                        idx = list(range(min(a, b), max(a, b) + 1))
                    else:
                        idx = [int(token)]
                    
                    idx = [i for i in idx if 1 <= i <= len(options)]
                    if not idx:
                        continue

                    grp = [options[i-1] for i in idx]
                    groups.append(grp)
            except Exception:
                print('Invalid entry, try again\n')
                continue
        
        if not groups:
            print('No valid selections, try again\n')
            continue

        print('\n')
        return groups


def _choose_plot():
    options = list(METRIC_CONFIG.keys())
    print('\n')

    return _choose('Select plot type:', options)


# -----------------------------
# Format Helpers
# -----------------------------
def _norm_dt(x):
    if isinstance(x, datetime):
        return x
    if isinstance(x, dtime):
        return datetime.combine(date.today(), x)
    if isinstance(x, str):
        s = x.strip()
        for fmt in ('%H:%M:%S.%f', '%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                t = datetime.strptime(s, fmt)
                
                if fmt in ('%H:%M:%S.%f', '%H:%M:%S'):
                    return datetime.combine(date.today(), t.time())
                return t
            except ValueError:
                pass

    raise TypeError(f'Unrecognized timestamp: {x!r}')


def _norm_float(x):
    try:
        return float(x.strip() if isinstance(x, str) else x)
    except Exception:
        return None


def _norm_date(df):
    s = str(df)
    for fmt in ('%m/%d/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return datetime.max


def _scale_feature(x):
    arr = np.asarray(x, dtype=float).ravel()

    finite_mask = np.isfinite(arr)
    if not finite_mask.any():
        return np.full_like(arr, np.nan, dtype=float)
    
    finite_vals = arr[finite_mask]
    
    mean = float(np.mean(finite_vals))
    std = float(np.std(finite_vals, ddof=0))

    x_scaled = np.full_like(arr, np.nan, dtype=float)

    if std <= 0.0 or not np.isfinite(std):
        x_scaled[finite_mask] = 0.0
    else:
        x_scaled[finite_mask] = (finite_vals - mean) / std
    
    return x_scaled


# -----------------------------
# Analysis Helpers
# -----------------------------
def _extract_config(session):
    th = session.get('threshold', None)
    bi = session.get('bidirectional', None)
    if (th is not None) and (bi is not None):
        return float(th), bool(bi)
    
    phase = session.get('phase')
    if isinstance(phase, str) and phase in PHASE_CONFIG:
        return PHASE_CONFIG[phase]
    
    return 30.0, True


def _extract_easy_rate(trials, hit_threshold=10):
    evt_list = [tr['evt'] for tr in trials][:19]

    K1 = 5
    n_hits = 0

    for evts in evt_list:
        if not evts:
            continue

        outcome = None
        for _, e in evts:
            if e in {'hit', 'miss'}:
                outcome = e
        if outcome == 'hit':
            n_hits += 1

    if n_hits < hit_threshold:
        K2 = 3
    elif n_hits == hit_threshold:
        K2 = 5
    else:
        K2 = 7

    return K1, K2


def _extract_trials(session):
    th, bi = _extract_config(session)

    evt_raw = session.get('EVT', []) or []
    enc_raw = session.get('ENC', []) or []

    evt = [
        (_norm_dt(t), str(e).strip().lower())
        for (t, e) in evt_raw
        if e is not None and str(e).strip() != ''
        ]
    enc = [
        (_norm_dt(t), _norm_float(d))
        for (t, d) in enc_raw
        ]
    
    if not evt or not enc:
        return []
    
    evt_nolick = [(t, e) for (t, e) in evt if e != 'lick']
    if len(evt_nolick) < 2:
        return []

    trials = []

    for r in range(0, len(evt_nolick) - 1, 2):
        t_cue, e_cue = evt_nolick[r]
        t_end, e_end = evt_nolick[r+1]

        if e_cue != 'cue':
            continue

        t0 = t_cue

        if r + 2 < len(evt_nolick):
            t_next_cue = evt_nolick[r+2][0]
        else:
            t_next_cue = None
        
        if t_next_cue is not None:
            evt_tr = [(t, e) for (t, e) in evt if t0 <= t < t_next_cue]
        else:
            evt_tr = [(t, e) for (t, e) in evt if t0 <= t]
        
        enc_tr = [(t, d) for (t, d) in enc if t0 < t <= t_end]

        if not evt_tr or not enc_tr:
            continue

        outcome_label = e_end if e_end in {'hit', 'miss'} else evt_tr[-1][1]

        trials.append({
            'enc': enc_tr,
            'evt': evt_tr,
            'outcome': outcome_label
            })
    
    return trials


def _extract_outcomes(trials):
    outcomes = []

    for tr in trials:
        if 'outcome' in tr and tr['outcome'] is not None:
            outcomes.append(tr['outcome'])
            continue

        outcome = None
        for _, e in tr.get('evt', []):
            if e in {'hit', 'miss'}:
                outcome = e

        if outcome is None and tr.get('evt'):
            outcome = tr['evt'][-1][1]

        outcomes.append(outcome)

    return outcomes


def _scalar(x):
    if np.isscalar(x):
        return float(x)
    
    arr = np.asarray(x, dtype=float).ravel()
    if arr.size != 1:
        raise ValueError(f'Expected a single value, got {arr.size}\n')
    
    return float(arr[0])


def _geometric_mean(arr):
        arr = np.array(arr, float)
        arr = arr[arr > 0]
        return float(np.exp(np.mean(np.log(arr)))) if len(arr) else np.nan


def _expected_value(arr, k=0.38):
    arr = np.array(arr, float)
    arr = arr[arr > 0]
    if arr.size == 0:
        return np.nan
    
    logs = np.log(arr)
    mu = np.mean(logs)
    var = np.var(logs, ddof=1)

    return float(np.exp(mu + (k * var)))


# -----------------------------
# Data Loading
# -----------------------------
def _scan_sessions(script_dir):
    meta = []
    col_idx = {}

    for filename in DATA_FILES:
        path = os.path.join(script_dir, filename)
        if not os.path.exists(path):
            continue

        wb = load_workbook(path,
                           read_only=True,
                           data_only=True
                           )
        if 'Event' not in wb.sheetnames or 'Encoder' not in wb.sheetnames:
            wb.close()
            continue

        evt_ws = wb['Event']
        enc_ws = wb['Encoder']

        per_file = {'Event': [], 'Encoder': []}

        enc_idx = {}
        for ec in range(1, enc_ws.max_column + 1, 2):
            d_v = enc_ws.cell(row=1, column=ec).value
            a_v = enc_ws.cell(row=2, column=ec).value
            p_v = enc_ws.cell(row=2, column=ec+1).value
            if d_v is None or a_v is None or p_v is None:
                continue
            d, a, p = str(d_v), str(a_v), str(p_v)

            if d and a and p:
                key = (a, p, d)
                enc_idx[key] = ec
                per_file['Encoder'].append({'date': d, 'animal': a, 'phase': p, 'col': ec})
        
        for c in range(1, evt_ws.max_column + 1, 2):
            sd_v  = evt_ws.cell(row=1, column=c).value
            sa_v = evt_ws.cell(row=2, column=c).value
            sp_v  = evt_ws.cell(row=2, column=c + 1).value
            if sd_v is None or sa_v is None or sp_v is None:
                continue
            s_date, s_animal, s_phase = str(sd_v), str(sa_v), str(sp_v)

            if not (s_date and s_animal and s_phase):
                continue
            if s_phase not in PHASE_CONFIG:
                continue

            enc_col = enc_idx.get((s_animal, s_phase, s_date))
            meta.append({
                'animal': s_animal,
                'phase':  s_phase,
                'date':   s_date,
                'file':   filename,
                'evt_col': c,
                'enc_col': enc_col
                })
            
            per_file['Event'].append({'date': s_date, 'animal': s_animal, 'phase': s_phase, 'col': c})
        
        col_idx[filename] = per_file
        wb.close()
    
    meta.sort(key=lambda m: (m['animal'], m['phase'], _norm_date(m['date'])))
    return meta


def _read_col_pairs(ws, cols):
    cols = sorted(set(c for c in cols if c is not None))
    if not cols:
        return {}

    data = {c: [] for c in cols}
    empty_streak = {c: 0 for c in cols}
    active = set(cols)

    min_col = min(cols)
    max_col = max(c + 1 for c in cols)

    for row in ws.iter_rows(min_row=4, min_col=min_col, max_col=max_col, values_only=True):
        if not active:
            break

        for c in list(active):
            ts_idx = c - min_col
            val_idx = ts_idx + 1

            if ts_idx >= len(row) or val_idx >= len(row):
                empty_streak[c] += 1
                if empty_streak[c] >= 3:
                    active.discard(c)
                continue

            ts = row[ts_idx]
            val = row[val_idx]

            if ts is None or ts == "":
                empty_streak[c] += 1
                if empty_streak[c] >= 3:
                    active.discard(c)
                continue

            empty_streak[c] = 0
            data[c].append([ts, val])

    return data


def _load_sessions(script_dir, metadata):
    N = len(metadata)
    print('Extracting session data...0%', end='', flush=True)

    by_file = {}
    for meta in metadata:
        by_file.setdefault(meta['file'], []).append(meta)

    data_map = {}
    processed = 0
    last_pct = 0

    for filename, group in by_file.items():
        path = os.path.join(script_dir, filename)
        if not os.path.exists(path):
            continue

        wb = load_workbook(path, read_only=True, data_only=True)
        if 'Event' not in wb.sheetnames or 'Encoder' not in wb.sheetnames:
            wb.close()
            continue

        evt_ws = wb['Event']
        enc_ws = wb['Encoder']

        evt_cols = sorted({m['evt_col'] for m in group if m.get('evt_col') is not None})
        enc_cols = sorted({m['enc_col'] for m in group if m.get('enc_col') is not None})

        evt_data_by_col = _read_col_pairs(evt_ws, evt_cols)
        enc_data_by_col = _read_col_pairs(enc_ws, enc_cols) if enc_cols else {}

        wb.close()

        for meta in group:
            key = (meta['animal'], meta['phase'], meta['date'])

            evt_data = evt_data_by_col.get(meta['evt_col'], [])
            enc_data = enc_data_by_col.get(meta['enc_col'], []) if meta.get('enc_col') is not None else []

            data_map[key] = {'EVT': evt_data, 'ENC': enc_data}

            processed += 1
            pct = int(100 * processed / N)
            if pct != last_pct:
                last_pct = pct
                print(f'\rExtracting session data...{pct}%', end='', flush=True)

            time.sleep(0.05)

    print('\rExtracting session data...100%\n', flush=True)

    return data_map


# -----------------------------
# Macros
# -----------------------------
DATA_FILES = (
    'EGI_data.xlsx',
    'QVWX_data.xlsx'
    )
PHASE_CONFIG = {
    'Phase 3': (30.0, True),
    'Phase 4': (60.0, True),
    'Phase 5': (90.0, True),
    }
METRIC_CONFIG = {
    'Percentage': {
        'key_fcn': _get_pct,
        'agg_fcn': [_scalar],
        'ylabel': ['Success %'],
        'ylim': [[(10, 110)]],
        'loc': 'lower left'
        },
    'Latency': {
        'key_fcn': _get_latency,
        'agg_fcn': [np.mean],
        'ylabel': ['Mean Trial Latency (s)'],
        'ylim': [[(0, 30)], [(0, 30)]]
        },
    'Displacement': {
        'key_fcn': _get_displacement,
        'agg_fcn': [_geometric_mean],
        'ylabel': ['Geometric Mean θ (°)'],
        'ylim': [[(0, 90), (0, 180)]]
        },
    'Velocity': {
        'key_fcn': _get_velocity,
        'agg_fcn': [_expected_value],
        'ylabel': [r'M[$\omega_{95\%}$] (RPM)'],
        'ylim': [[(0, 600)]]
        },
    'Acceleration': {
        'key_fcn': _get_acceleration,
        'agg_fcn': [_expected_value],
        'ylabel': [r'M[$\alpha_{95\%}$] (rev/min$^{2}$)'],
        'ylim': [[(0, 1.2e6)]]
        },
    'Duration': {
        'key_fcn': _get_duration,
        'agg_fcn': [_scalar],
        'ylabel': ['Session Duration (min)'],
        'ylim': [[(0, 50)]]
        },
    'Licks': {
        'key_fcn': _get_licks,
        'agg_fcn': [np.mean],
        'ylabel': ['Relative Mean Lick Time (s)'],
        'ylim': [[(-30, 3)]]
    }
    }

HIST_ANIMAL = 'Animal G'
HIST_DATE = '11/18/2025'

# -----------------------------
# Diffusion-Map Pipeline
# -----------------------------
def _get_all_features():
    names = []
    for metric in METRIC_CONFIG:
        if metric == 'Duration':
            names.append('Duration')
        else:
            names.append(f'{metric} (easy)')
            names.append(f'{metric} (normal)')
            names.append(f'{metric} (all)')
    
    return names


def _build_feature_matrix(script_dir, selected_features=None):
    all_meta = _scan_sessions(script_dir)
    if not all_meta:
        print('No session data found')
        return None, None, None, None
    
    meta = [m for m in all_meta if m['phase'] in PHASE_CONFIG]
    if not meta:
        print('No sessions in configured phases')
        return None, None, None, None
    
    data_map = _load_sessions(script_dir, meta)

    feature_names = _get_all_features()    
    n_features = len(feature_names)

    feature_acc = {}
    for m in meta:
        animal = m['animal']
        phase = m['phase']
        date_str = m['date']

        session = data_map.get((animal, phase, date_str))
        if not session:
            continue

        th, bi = PHASE_CONFIG[phase]
        cfg_s = dict(session)
        cfg_s['phase'] = phase
        cfg_s['threshold'] = th
        cfg_s['bidirectional'] = bi

        trials = _extract_trials(cfg_s)
        if not trials:
            continue

        outcomes = _extract_outcomes(trials)
        if not outcomes:
            continue

        session_dt = _norm_date(date_str)
        hit_threshold = 16 if (session_dt.month == 11 and session_dt.day in {11, 12}) else 10
        K1, K2 = _extract_easy_rate(trials, hit_threshold)
        stop_idx = _filter_trials((K1, K2), outcomes)

        trials = trials[:stop_idx]
        outcomes = outcomes[:stop_idx]
        if not trials:
            continue

        t_start = trials[0]['evt'][0][0]
        t_stop = trials[-1]['evt'][-1][0]
        
        for tr in trials:
            tr['elapsed'] = (t_stop - t_start).total_seconds() / 60.0
        
        is_easy = []
        for i in range(1, len(trials) + 1):
            if i <= 20:
                easy = ((i - 1) % K1 == 0)
            else:
                easy = ((i - 21) % K2 == 0)
            
            is_easy.append(easy)
        
        idx_easy = [i for (i, e) in enumerate(is_easy) if e]
        idx_normal = [i for (i, e) in enumerate(is_easy) if not e]

        sample_key = (animal, date_str)
        feature_acc.setdefault(sample_key, {})

        for metric, cfg in METRIC_CONFIG.items():
            key_fcn = cfg['key_fcn']
            agg_fcn = cfg['agg_fcn'][0]

            if metric == 'Duration':
                vals = key_fcn(trials=trials)
                if vals is None:
                    continue

                vals = np.array(vals if not np.isscalar(vals) else [vals], dtype=float)
                if vals.size == 0:
                    continue

                try:
                    mv = float(agg_fcn(vals))
                except Exception:
                    continue

                feature_acc[sample_key].setdefault('Duration', []).append(mv)
            else:
                if idx_easy:
                    t_easy = [trials[i] for i in idx_easy]
                    o_easy = [outcomes[i] for i in idx_easy]

                    vals_e = key_fcn(t_easy, o_easy)

                    if vals_e is not None:
                        vals_e = np.array(vals_e if not np.isscalar(vals_e) else [vals_e], dtype=float)

                        if vals_e.size:
                            try:
                                mv = float(agg_fcn(vals_e))
                                fname = f'{metric} (easy)'
                                
                                feature_acc[sample_key].setdefault(fname, []).append(mv)
                            except Exception:
                                pass
                
                if idx_normal:
                    t_normal = [trials[i] for i in idx_normal]
                    o_normal = [outcomes[i] for i in idx_normal]

                    vals_n = key_fcn(t_normal, o_normal)
                    
                    if vals_n is not None:
                        vals_n = np.array(vals_n if not np.isscalar(vals_n) else [vals_n], dtype=float)

                        if vals_n.size:
                            try:
                                mv = float(agg_fcn(vals_n))
                                fname = f'{metric} (normal)'

                                feature_acc[sample_key].setdefault(fname, []).append(mv)
                            except Exception:
                                pass
                
                vals_all = key_fcn(trials, outcomes)
                if vals_all is not None:
                    vals_all = np.array(vals_all if not np.isscalar(vals_all) else [vals_all], dtype=float)

                    if vals_all.size:
                        try:
                            mv = float(agg_fcn(vals_all))
                            fname = f'{metric} (all)'

                            feature_acc[sample_key].setdefault(fname, []).append(mv)
                        except Exception:
                            pass

    if not feature_acc:
        print('No usable feature data')
        return None, None, None, None
    
    sample_keys = sorted(feature_acc.keys(), key=lambda k: (_norm_date(k[1]), k[0]))
    n_samples = len(sample_keys)

    X_raw = np.full((n_samples, n_features), np.nan, dtype=float)

    for i, key in enumerate(sample_keys):
        feature = feature_acc[key]

        for j, fname in enumerate(feature_names):
            vals = feature.get(fname)
            if vals:
                X_raw[i, j] = float(np.mean(vals))
    
    X_scaled = X_raw.copy()
    
    for j in range(n_features):
        col = _scale_feature(X_scaled[:, j])
        
        mask = np.isfinite(col)
        if mask.any():
            m = float(np.mean(col[mask]))
        else:
            m = 0.0
        
        col[~mask] = m
        X_scaled[:, j] = col
    
    unique_dates = sorted({k[1] for k in sample_keys}, key=_norm_date)
    date_to_idx = {d: i for (i, d) in enumerate(unique_dates)}

    sample_idx = np.array([date_to_idx[k[1]] for k in sample_keys], dtype=int)

    if selected_features:
        selected_set = set(selected_features)
        keep_idx = [j for (j, name) in enumerate(feature_names) if name in selected_set]

        if keep_idx:
            X_raw = X_raw[:, keep_idx]
            X_scaled = X_scaled[:, keep_idx]
            feature_names = [feature_names[j] for j in keep_idx]

    return X_raw, X_scaled, sample_keys, sample_idx, feature_names


def _diffusion_map(X, epsilon=None):
    if X is None or X.size == 0:
        return None
    
    D = squareform(pdist(X, metric='euclidean'))

    if epsilon is None:
        dvals = D[np.triu_indices_from(D, k=1)]
        dvals = dvals[dvals > 0]

        if dvals.size == 0:
            epsilon = 1.0
        else:
            med = np.median(dvals)
            epsilon = med ** 2
    
    K = np.exp(-(D ** 2) / epsilon)

    row_sums = K.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1.0
    
    P = K / row_sums

    evals, evecs = np.linalg.eig(P.T)

    idx = np.argsort(-evals.real)
    evals = evals[idx].real
    evecs = evecs[:, idx].real

    if evals.size < 2:
        return None
    
    psi = evecs[:, 1]
    psi = (psi - np.mean(psi)) / (np.std(psi) if np.std(psi) > 0 else 1.0)

    return psi


def _analyze_features(X, psi, feature_names):
    X = np.asarray(X, float)
    psi = np.asarray(psi, float).ravel()
    n_samples, n_features = X.shape

    if psi.size != n_samples:
        print('Shape mismatch between X and psi')
        return
    
    corrs = np.full(n_features, np.nan, dtype=float)
    
    for j in range(n_features):
        xj = X[:, j]
        mask = np.isfinite(xj) & np.isfinite(psi)
        if mask.sum() < 2:
            continue

        xj_c = xj[mask] - xj[mask].mean()
        psi_c = psi[mask] - psi[mask].mean()

        denom = (np.linalg.norm(xj_c) * np.linalg.norm(psi_c))
        if denom > 0:
            corrs[j] = float(np.dot(xj_c, psi_c) / denom)
    
    try:
        beta, _, _, _ = np.linalg.lstsq(X, psi, rcond=None)
    except Exception:
        beta = np.full(n_features, np.nan, dtype=float)
    
    order = np.argsort(-np.abs(beta))

    print(f'{"Rank":>4s}  {"Feature":30s}  {"Weight":>10s}  {"Corr":>10s}')
    print('-' * 62)

    for rank, j in enumerate(order, start=1):
        name = feature_names[j] if j < len(feature_names) else f'feat_{j}'
        w = beta[j]
        c = corrs[j]
        w_str = f'{w:+.4f}' if np.isfinite(w) else '  nan'
        c_str = f'{c:+.4f}' if np.isfinite(c) else '  nan'

        print(f'{rank:4d}  {name:30s}  {w_str:>10s}  {c_str:>10s}')
    
    print('\n')


def _plot_corr(X, psi, feature_names, sample_keys, raw_X=None, win_title=''):
    """
    Create a subplot grid where each subplot shows
    a single metric (x-axis) vs the scalar values (y-axis).
    """
    X = np.asarray(X, float)
    psi = np.asarray(psi, float).ravel()

    n_samples, n_features = X.shape
    if psi.size != n_samples:
        print('Shape mismatch between X and psi in _plot_corr')
        return

    if n_features == 0:
        print('No features to plot in _plot_corr')
        return

    # Grid layout
    n_rows = min(2, n_features)  # up to 3 columns
    n_cols = int(np.ceil(n_features / n_rows))

    fig, axes = plt.subplots(
        n_rows,
        n_cols,
        figsize=(4.0 * n_cols, 3.0 * n_rows),
        squeeze=False
    )

    axes_flat = axes.ravel()

    for j in range(n_features):
        ax = axes_flat[j]

        xj = X[:, j]
        mask = np.isfinite(xj) & np.isfinite(psi)
        if mask.sum() < 2:
            ax.set_visible(False)
            continue

        xj = xj[mask]
        yj = psi[mask]

        yj = [-y for y in yj]
        yj = [y - np.min(yj) for y in yj]

        ax.scatter(xj, yj, s=25, color='red', edgecolors='none')
        ax.set_title(
            feature_names[j] if j < len(feature_names) else f'Feature {j}',
            fontsize=11,
            fontweight='bold'
        )
        ax.set_xlabel('Feature Value', fontsize=9)
        ax.set_ylabel('\u03A8', fontsize=9)  # using the same sign convention as main plot
        ax.grid(True, alpha=0.3)

    # Hide any unused axes
    for k in range(n_features, len(axes_flat)):
        axes_flat[k].set_visible(False)

    fig.tight_layout()
    show_figure(fig, win_title)

    if raw_X is not None and sample_keys is not None:
        raw_X = np.asarray(raw_X, dtype=float)

        if raw_X.shape == X.shape and len(sample_keys) == n_samples:
            samples_out = []

            for i in range(n_samples):
                animal, date_str = sample_keys[i]
                feature_dict = {}

                for j, fname in enumerate(feature_names):
                    val = raw_X[i, j]

                    if np.isnan(val):
                        feature_dict[fname] = None
                    else:
                        feature_dict[fname] = float(val)
                
                psi_val = psi[i]
                psi_val = float(psi_val) if np.isfinite(psi_val) else None

                samples_out.append({
                    'animal': animal,
                    'date': date_str,
                    'psi': psi_val,
                    'features': feature_dict
                    })
            
            corr_data = {'samples': samples_out}

            script_dir = os.path.dirname(os.path.abspath(__file__))
            out_path = os.path.join(script_dir, 'correlation_data.json')

            with open(out_path, 'w') as f:
                json.dump(corr_data, f, indent=2)


def run_pipeline():
    '''
    1) Let user select one or more features to include
    2) Build feature matrix using selected metrics and all session dates
    3) Run diffusion-map (1D) to get latent scalar per (animal, date) pair
    4) Shift all animal series by the global minimum so values are >= 0
    5) Compute geometric mean across animals per session index
    6) Plot per-animal shifted series and the geometric-mean series
    '''
    # --- Feature selection ---
    all_feature_names = _get_all_features()
    metric_groups = _choose_multi('Select one or more features:', all_feature_names)

    selected_metrics = []
    for grp in metric_groups:
        for name in grp:
            if name not in selected_metrics:
                selected_metrics.append(name)
    
    if not selected_metrics:
        selected_metrics = all_feature_names

    print('\nRunning diffusion-map manifold analysis with features:')
    for name in selected_metrics:
        print(f'  - {name}')
    print()

    # --- Build feature matrix ---
    script_dir = os.path.dirname(os.path.abspath(__file__))
    X_raw, X_scaled, sample_keys, sample_idx, feature_names = _build_feature_matrix(
        script_dir,
        selected_features=selected_metrics
    )
    
    if X_scaled is None:
        print('Diffusion pipeline aborted (no data)')
        return
    
    # --- Diffusion map ---
    psi = _diffusion_map(X_scaled)
    if psi is None:
        print('Diffusion map failed')
        return
    
    # Feature contribution analysis
    _analyze_features(X_scaled, psi, feature_names)

    psi = np.asarray(psi, float)

    # Performance-oriented coordinate: higher = better
    psi_perf = -psi

    # Global shift: subtract the absolute minimum across all animals/days
    global_min = float(np.min(psi_perf))
    phi_shift = psi_perf - global_min  # at least one value is exactly 0, others >= 0

    # --- Index and date bookkeeping ---
    sample_idx_arr = np.asarray(sample_idx, dtype=int)
    unique_idx = sorted(set(sample_idx_arr.tolist()))

    idx_to_dates = {}
    for i, (_, date_str) in enumerate(sample_keys):
        di = sample_idx_arr[i]
        idx_to_dates.setdefault(di, set()).add(date_str)
    
    # --- Geometric mean across animals per session index (day) ---
    latent_idx = {}
    for di in unique_idx:
        mask = (sample_idx_arr == di)
        vals = phi_shift[mask]  # shifted, >= 0

        if vals.size:
            latent_idx[di] = _geometric_mean(vals)

    items = sorted(latent_idx.items(), key=lambda kv: kv[0])
    if not items:
        print('No latent values computed')
        return
    
    idx = [i for (i, _) in items]
    values = [v for (_, v) in items]

    # --- Print summary of geometric-mean performance coordinate ---
    print('\nsample index -> geometric-mean performance coordinate (φ_shift):')
    print('-' * 70)
    for di, val in items:
        dates = sorted(idx_to_dates.get(di, []), key=_norm_date)
        dstr = ', '.join(dates)
        print(f'\t[{dstr}]\t{di:2d}: {val:.4f}')

    # --- Build per-animal shifted series ---
    animal_series = {}
    for i, (animal, date_str) in enumerate(sample_keys):
        di = sample_idx_arr[i]
        y_val = phi_shift[i]  # shifted performance coordinate for this animal/date

        animal_series.setdefault(animal, []).append((di, y_val))
    
    # Sort each animal's points by session index
    for animal in animal_series:
        animal_series[animal].sort(key=lambda p: p[0])
    
    # --- Save performance line-plot data as JSON ---
    performance_data = {
        'feature_set': [str(fname) for fname in feature_names],
        'session_index_to_dates': {
            str(di): sorted(list(dates), key=_norm_date)
            for (di, dates) in idx_to_dates.items()
            },
        'geometric_mean': {
            'session_indices': [int(i) for i in idx],
            'phi': [float(v) for v in values]
            },
        'animals': {
            animal: {
                'session_indices': [int(p[0]) for p in pts],
                'phi': [float(p[1]) for p in pts]
                }
            for (animal, pts) in animal_series.items()
            }
        }
    
    perf_json_path = os.path.join(script_dir, 'performance_data.json')
    
    with open(perf_json_path, 'w') as f:
        json.dump(performance_data, f, indent=2)
    
    # --- Plotting ---
    fig, ax = plt.subplots(figsize=(8, 5))

    # 1) Individual animal trajectories
    for animal, pts in animal_series.items():
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]

        ax.plot(xs, ys,
                linewidth=1.5,
                marker='o',
                markersize=6,
                label=str(animal)
                )

    # 2) Geometric-mean trajectory (black line)
    ax.plot(idx, values,
            linewidth=2.0,
            color='black',
            marker='o',
            markersize=7,
            label='Geometric Mean'
            )
    
    feature_set_combo = 'C'
    
    ax.set_xlabel('Training Day',
                  fontsize=13,
                  fontweight='bold'
                  )
    ax.set_ylabel('Latent Scalar (\u03C6)',
                  fontsize=13,
                  fontweight='bold'
                  )
    ax.set_title(f'Feature Set {feature_set_combo}',
                 fontsize=16,
                 fontweight='bold'
                 )
    
    # X-ticks labelled by dates
    xticks = idx
    ax.set_xticklabels(range(1, len(xticks) + 1))

    ax.grid(True, alpha=0.3)
    ax.legend(loc='best', fontsize=9)
    
    fig.tight_layout()
    show_figure(fig)

    # plt.savefig(os.path.join(script_dir, f'map_{feature_set_combo}.png'), dpi=600)

    # 3) Feature correlations with latent scalar values
    _plot_corr(X_scaled, phi_shift, feature_names, sample_keys, raw_X=X_raw)
    # plt.savefig(os.path.join(script_dir, f'corr_{feature_set_combo}.png'), dpi=600)


# -----------------------------
# Plotting
# -----------------------------
def _hist(trials, outcomes, key_fcn):
    vals = key_fcn(trials, outcomes)
    if vals is None:
        return

    vals = np.asarray(vals if not np.isscalar(vals) else [vals], float)
    vals = vals[np.isfinite(vals)]
    if vals.size == 0:
        return

    data_min = float(np.min(vals))
    data_max = float(np.max(vals))
    data_range = data_max - data_min
    if data_range <= 0:
        data_range = 1.0

    q25, q75 = np.percentile(vals, [25, 75])
    iqr = q75 - q25
    n = len(vals)

    if iqr > 0 and n > 1:
        h = 2 * iqr / (n ** (1/3))
    else:
        h = data_range / 7.0

    bins = max(1, int(np.ceil(data_range / h)))

    counts, edges = np.histogram(vals,
                                 bins=bins,
                                 range=(data_min, data_max)
                                 )
    centers = 0.5 * (edges[:-1] + edges[1:])

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(centers, counts,
           width=np.diff(edges),
           align='center',
           edgecolor='black'
           )

    ax.grid(True, alpha=0.3)

    title = '<missing>'
    for metric, cfg in METRIC_CONFIG.items():
        if cfg['key_fcn'] is key_fcn:
            title = metric
            break
    
    ax.set_title(title,
                 fontsize=14,
                 fontweight='bold'
                 )

    fig.tight_layout()
    show_figure(fig)


def plot_sessions(ax, x, y, marker, label):
    if not x or not y:
        return
    
    x_arr = np.asarray(x, dtype=float)
    y_arr = np.asarray(y, dtype=float)

    mask = ~np.isnan(x_arr) & ~np.isnan(y_arr)
    if not mask.any():
        return
    
    x_arr = x_arr[mask]
    y_arr = y_arr[mask]

    if x_arr.size  == 0:
        return
    
    ax.plot(x_arr, y_arr,
            linewidth=1.5,
            marker=marker,
            markersize=6,
            label=str(label)
            )


def plot_trial(script_dir, data_map):
    sessions = list(data_map.keys())
    if not sessions:
        print("No sessions available")
        return

    best_jaggedness = None
    best_choice = None  # (s_key, idx, trials, th)

    WINDOW = 5.0  # seconds
    STEP = 0.5    # seconds

    for s_key in sessions:
        session = data_map[s_key]
        _, phase, _ = s_key

        th, bi = PHASE_CONFIG.get(phase, (30.0, True))

        cfg_s = dict(session)
        cfg_s["phase"] = phase
        cfg_s["threshold"] = th
        cfg_s["bidirectional"] = bi

        trials = _extract_trials(cfg_s)
        if not trials:
            continue

        outcomes = _extract_outcomes(trials)
        if not outcomes:
            continue

        # Only 'miss' trials
        miss_idx = [i for (i, o) in enumerate(outcomes) if o == "miss"]
        if not miss_idx:
            continue

        for idx in miss_idx:
            tr = trials[idx]
            enc = tr.get("enc", [])
            if not enc:
                continue

            enc_valid = [(t, v) for (t, v) in enc if v is not None]
            if len(enc_valid) < 2:
                continue

            # Raw arrays
            times_raw = [t for (t, _) in enc_valid]
            values_raw = np.array([v for (_, v) in enc_valid], dtype=float)

            # Interpolate using your existing logic
            times, values = _interp(times_raw, values_raw)
            if len(times) < 2:
                continue

            # Relative times in seconds
            t0 = times[0]
            ts = np.array([(t - t0).total_seconds() for t in times], dtype=float)

            T = ts[-1]
            if T < WINDOW:
                # Not enough duration for even one 5-s window
                continue

            # Compute 5-s windows stepping by 1 s
            win_vars = []
            start = 0.0
            while start + WINDOW <= T + 1e-9:
                end = start + WINDOW
                mask = (ts >= start) & (ts <= end)
                v_win = values[mask]

                if v_win.size > 1:
                    win_vars.append(float(np.var(v_win)))

                start += STEP

            if not win_vars:
                continue

            # Jaggedness metric that rewards sustained variability:
            # median variance across all windows
            jaggedness = float(np.percentile(win_vars, 25))
            # Keep the single most jagged (sustained) trial
            if (best_jaggedness is None) or (jaggedness > best_jaggedness):
                best_jaggedness = jaggedness
                best_choice = (s_key, idx, trials, th)

    if best_choice is None:
        print("No valid 'miss' trials found")
        return

    # Unpack most jagged trial
    s_key, idx, trials, th = best_choice
    tr = trials[idx]

    enc_valid = [(t, v) for (t, v) in tr["enc"] if v is not None]
    if len(enc_valid) < 2:
        print("No valid displacement samples for chosen trial")
        return

    times = [t for (t, _) in enc_valid]
    values = np.array([v for (_, v) in enc_valid], dtype=float)

    # Relative time
    t0 = times[0]
    t_rel = np.array([(t - t0).total_seconds() for t in times], dtype=float)

    # Extend visually to 30 s, if you still want that behavior
    if t_rel[-1] < 30.0:
        t_rel = np.append(t_rel, 30.0)
        values = np.append(values, values[-1])

    # Plot
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(t_rel, values, linewidth=1.2, color="blue")

    ax.set_xlim(0, 30)
    ax.set_ylim(-th - 5, th + 5)

    ax.set_yticks(np.arange(-th, th + 1, 10))
    ax.axhline(th, color="black", linestyle=(6, (6, 6)), linewidth=1.0)
    ax.axhline(-th, color="black", linestyle=(6, (6, 6)), linewidth=1.0)

    ax.set_xlabel("Time [sec]", fontsize=12)
    ax.set_ylabel("Wheel Displacement [°]", fontsize=12)
    ax.set_title(
        f"Most Jagged 'Miss' Trial\n(max median 5-s variance = {best_jaggedness:.2f})"
    )

    fig.tight_layout()
    show_figure(fig)

    save_opt = input("\nSave figure? [y/N]: ").strip().lower()
    if save_opt == "y":
        out_csv = os.path.join(script_dir, 'disp.csv')

        with open(out_csv, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['time_seconds', 'displacement'])

            for t, v in zip(t_rel, values):
                writer.writerow([t, v])

    print()


def show_figure(fig, win_title=''):
    buf = BytesIO()
    fig.savefig(buf,
                format='png',
                bbox_inches='tight'
                )
    buf.seek(0)

    pg.init()

    try:
        desktop_sizes = pg.display.get_desktop_sizes()
    except AttributeError:
        info = pg.display.Info()
        desktop_sizes = [(info.current_w, info.current_h)]
    
    screen_size = desktop_sizes[0]

    if len(desktop_sizes) >= 2:
        primary_w, _ = desktop_sizes[0]
        secondary_w, secondary_h = desktop_sizes[1]
    
        os.environ['SDL_VIDEO_WINDOW_POS'] = f'{primary_w},0'
        screen_size = (secondary_w, secondary_h)
    else:
        os.environ.setdefault('SDL_VIDEO_CENTERED', '1')
    
    screen = pg.display.set_mode(screen_size)
    pg.display.set_caption(win_title)

    plot_surface = pg.image.load(buf)

    img_w, img_h = plot_surface.get_size()
    win_w, win_h = screen_size
    scale = min(win_w / img_w, win_h / img_h)
    new_size = (int(img_w * scale), int(img_h * scale))
    plot_surface = pg.transform.smoothscale(plot_surface, new_size)

    rect = plot_surface.get_rect(center=(win_w // 2, win_h // 2))

    clock = pg.time.Clock()
    running = True

    while running:
        for event in pg.event.get():
            if event.type == pg.QUIT:
                running = False
            elif event.type == pg.KEYDOWN and event.key == pg.K_ESCAPE:
                running = False
        
        screen.fill((0, 0, 0))
        screen.blit(plot_surface, rect)
        pg.display.flip()
        clock.tick(60)
    
    pg.quit()

    save_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '__figures__')
    os.makedirs(save_dir, exist_ok=True)

    out_path = os.path.join(save_dir, f'{win_title}.png')
    fig.savefig(out_path, dpi=300, bbox_inches='tight')


def main():
    print('\nScanning workbooks...\n\n')

    script_dir = os.path.dirname(os.path.abspath(__file__))
    all_meta = _scan_sessions(script_dir)
    if not all_meta:
        print('No session data found')
        return
    
    data_map = _load_sessions(script_dir, all_meta)

    plot_trial(script_dir, data_map)
    exit(0)
    
    all_meta = [m for m in all_meta if m['phase'] in PHASE_CONFIG]

    all_dates = sorted({m['date'] for m in all_meta}, key=_norm_date)
    if not all_dates:
        print('No dates found')
        return
    
    date_groups = _choose_multi('Select one or more session dates:', all_dates)
    selected_dates = sorted({d for group in date_groups for d in group},
                            key=_norm_date)

    meta = [m for m in all_meta if m['date'] in set(selected_dates)]
    if not meta:
        print('No sessions found in selected date range')
        return
    
    date_to_group_idx = {}
    for gi, group in enumerate(date_groups):
        for d in group:
            date_to_group_idx[d] = gi
    
    group_labels = []
    for group in date_groups:
        if not group:
            group_labels.append('')
            continue
        
        group_sorted = sorted(group, key=_norm_date)

        start = _norm_date(group_sorted[0])
        stop = _norm_date(group_sorted[-1])
        if start == stop:
            label = start.strftime('%m/%d')
        else:
            label = f"{start.strftime('%m/%d')} - {stop.strftime('%m/%d')}"
        
        group_labels.append(label)

    plot_type = _choose_plot()

    cfg = METRIC_CONFIG[plot_type]
    key_fcn = cfg['key_fcn']
    agg_fcns = (cfg['agg_fcn']
                if isinstance(cfg['agg_fcn'], list)
                else [cfg['agg_fcn']])
    ylabels = (cfg['ylabel']
               if isinstance(cfg['ylabel'], list)
               else [cfg['ylabel']])
    ylims = cfg['ylim']
    loc = cfg.get('loc', 'best')

    n_agg = len(agg_fcns)

    if plot_type == 'Duration':
        fig, axes = plt.subplots(n_agg, 1, figsize=(8, 5 * n_agg))
    else:
        fig, axes = plt.subplots(n_agg, 2, figsize=(12, 5 * n_agg))
    
    axes = np.atleast_2d(axes)
    
    data_map = _load_sessions(script_dir, meta)
    animals = sorted({m['animal'] for m in meta})

    ##################
    def _compute_animal_series(animal, mode, agg_fcn, per_group_acc=None):
        group_values = {}

        for m in [mm for mm in meta if mm['animal'] == animal]:
            session = data_map.get((m['animal'], m['phase'], m['date']))
            if not session:
                continue

            th, bi = PHASE_CONFIG[m['phase']]

            cfg_s = dict(session)
            cfg_s['phase'] = m['phase']
            cfg_s['threshold'] = th
            cfg_s['bidirectional'] = bi

            trials = _extract_trials(cfg_s)
            if not trials:
                continue

            outcomes = _extract_outcomes(trials)
            if not outcomes:
                continue

            session_date = _norm_date(m['date'])
            hit_threshold = 16 if (session_date.month == 11 and session_date.day in {11, 12}) else 10
            K1, K2 = _extract_easy_rate(trials, hit_threshold)
            stop_idx = _filter_trials((K1, K2), outcomes)

            trials = trials[:stop_idx]
            outcomes = outcomes[:stop_idx]

            if trials:
                t_start = trials[0]['evt'][0][0]
                t_stop = trials[-1]['evt'][-1][0]

                for tr in trials:
                    tr['elapsed'] = (t_stop - t_start).total_seconds() / 60.0

            is_easy = []
            for i in range(1, len(trials) + 1):
                if i <= 20:
                    easy = ((i - 1) % K1 == 0)
                else:
                    easy = ((i - 21) % K2 == 0)
                
                is_easy.append(easy)

            if mode == 'easy':
                idx = [i for (i, e) in enumerate(is_easy) if e]
            else:
                idx = [i for (i, e) in enumerate(is_easy) if not e]
            
            if not idx:
                continue

            vals = key_fcn([trials[i] for i in idx], [outcomes[i] for i in idx])
            if vals is None:
                continue

            vals = np.array(vals if not np.isscalar(vals) else [vals], dtype=float)
            if len(vals) == 0:
                continue

            mv = float(agg_fcn(vals))
            gi = date_to_group_idx[m['date']]

            group_values.setdefault(gi, []).append(mv)

        if not group_values:
            return [], []
        
        per_group = {gi: float(np.mean(vs)) for (gi, vs) in group_values.items() if vs}
        if not per_group:
            return [], []
        
        if per_group_acc is not None:
            for gi, mv in per_group.items():
                per_group_acc.setdefault(gi, []).append(mv)
        
        xs = sorted(per_group.keys())
        ys = [per_group[x] for x in xs]

        return xs, ys
    ##################

    for fi, agg_fcn in enumerate(agg_fcns):
        yl = ylabels[fi]
        ylim_list = ylims[fi]

        if plot_type == 'Duration':
            ax = axes[fi, 0]

            markers = cycle(['D', 'X', 's', 'o', '^'])
            geo_acc = {}

            for animal in animals:
                marker = next(markers)
                xs, ys = _compute_animal_series(animal, 'normal', agg_fcn, geo_acc)

                if xs:
                    plot_sessions(ax, xs, ys, marker, animal)
            
            xs_geo = []
            ys_geo = []

            for x in sorted(geo_acc.keys()):
                val = _geometric_mean(geo_acc[x])

                if not np.isnan(val):
                    xs_geo.append(x)
                    ys_geo.append(val)
            
            if not xs_geo:
                continue

            ax.plot(xs_geo, ys_geo,
                    linewidth=1.5,
                    marker='o',
                    markersize=6,
                    color='black',
                    label='Geometric Mean'
                    )
            
            ax.set_ylim(*ylim_list[0])
            ax.set_xlim(-0.5, len(group_labels) - 0.5)
            ax.set_xticks(range(len(group_labels)))
            ax.set_xticklabels(group_labels, rotation=45)

            ax.set_xlabel('Date',
                          fontsize=13,
                          fontweight='bold'
                          )
            ax.set_ylabel(yl,
                          fontsize=13,
                          fontweight='bold'
                          )
            
            ax.grid(True, alpha=0.3)
            ax.legend(loc=loc, fontsize=7.5)

            if len(xs_geo) >= 2:
                m, _ = np.polyfit(xs_geo, ys_geo, deg=1)
                ax.set_title(f'Duration (m = {m:.2f})',
                                fontsize=14,
                                fontweight='bold'
                                )
            else:
                ax.set_title('Duration',
                             fontsize=14,
                             fontweight='bold'
                             )
            
            continue
        else:
            easy_ax = axes[fi, 0]
            normal_ax = axes[fi, 1]

            if len(ylim_list) == 1:
                easy_lim = ylim_list[0]
                normal_lim = ylim_list[0]
            else:
                easy_lim, normal_lim = ylim_list

            markers = cycle(['D', 'X', 's', 'o', '^'])

            geo_easy = {}
            geo_normal = {}

            for animal in animals:
                marker = next(markers)

                xe, ye = _compute_animal_series(animal, 'easy', agg_fcn, geo_easy)
                if xe:
                    plot_sessions(easy_ax, xe, ye, marker, animal)
                
                xn, yn = _compute_animal_series(animal, 'normal', agg_fcn, geo_normal)
                if xn:
                    plot_sessions(normal_ax, xn, yn, marker, animal)
            
            xs_e = []
            ys_e = []

            for x in sorted(geo_easy.keys()):
                val = _geometric_mean(geo_easy[x])
                if not np.isnan(val):
                    xs_e.append(x)
                    ys_e.append(val)
            
            xs_n = []
            ys_n = []
            for x in sorted(geo_normal.keys()):
                val = _geometric_mean(geo_normal[x])
                if not np.isnan(val):
                    xs_n.append(x)
                    ys_n.append(val)

            if xs_e:
                easy_ax.plot(xs_e, ys_e,
                            linewidth=1.5,
                            marker='o',
                            markersize=6,
                            color='black',
                            label='Geometric Mean'
                            )
            if xs_n:
                normal_ax.plot(xs_n, ys_n,
                            linewidth=1.5,
                            marker='o',
                            markersize=6,
                            color='black',
                            label='Geometric Mean'
                            )
            
            easy_ax.set_ylim(*easy_lim)
            normal_ax.set_ylim(*normal_lim)

            n_groups = len(group_labels)

            for ax in (easy_ax, normal_ax):
                ax.set_xlim(-0.5, n_groups - 0.5)
                ax.set_xticks(range(n_groups))
                ax.set_xticklabels(group_labels, rotation=45)
                ax.tick_params(
                    axis='both',
                    labelsize=12,
                    labelcolor='black'
                    )

                ax.set_xlabel('Date',
                            fontsize=13,
                            fontweight='bold'
                            )
                ax.set_ylabel(yl,
                            fontsize=13,
                            fontweight='bold'
                            )

                ax.grid(True, alpha=0.3)
                ax.legend(loc=loc, fontsize=7.5)
            
            if len(xs_e) >= 2:
                m_e, _ = np.polyfit(xs_e, ys_e, deg=1)
                easy_ax.set_title(f'Easy Trials (m = {m_e:.2f})',
                                  fontsize=14,
                                  fontweight='bold'
                                  )
            else:
                easy_ax.set_title('Easy Trials',
                                  fontsize=14,
                                  fontweight='bold'
                                  )
            
            if len(xs_n) >= 2:
                m_n, _ = np.polyfit(xs_n, ys_n, deg=1)
                normal_ax.set_title(f'Normal Trials (m = {m_n:.2f})',
                                    fontsize=14,
                                    fontweight='bold'
                                    )
            else:
                normal_ax.set_title('Normal Trials',
                                    fontsize=14,
                                    fontweight='bold'
                                    )
    
    fig.tight_layout(pad=2.0)
    show_figure(fig, plot_type)

    save_opt = input('\nSave figure? [y/N]:  ').strip().lower()
    save_opt = {'y': True, 'n': False}[save_opt]

    if save_opt:
        plt.savefig(os.path.join(script_dir, f'{plot_type}.png'), dpi=600)
    
    print()


def plot_success():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    all_meta = _scan_sessions(script_dir)
    meta = [m for m in all_meta if m['file'] == 'QVWX_data.xlsx' and m['phase'] in PHASE_CONFIG]

    if not meta:
        print('No sessions found in QVWX_data.xlsx')
        return
    
    data_map = _load_sessions(script_dir, meta)
    
    animals = sorted({m['animal'] for m in meta})
    if not animals:
        print('No animals found in QVWX_data.xlsx')
        return
    
    n_animals = len(animals)
    n_cols = 2
    n_rows = 2

    fig, axes = plt.subplots(n_rows, n_cols,
                             figsize=(8, 5),
                             squeeze=False
                             )
    axes_flat = axes.ravel()

    for ax in axes_flat:
        ax.set_visible(False)

    for idx_animal, animal in enumerate(animals):
        ax = axes_flat[idx_animal]
        ax.set_visible(True)

        meta_a = [m for m in meta if m['animal'] == animal]
        meta_a.sort(key=lambda m: _norm_date(m['date']))

        for m in meta_a:
            phase = m['phase']
            date_str = m['date']

            session_key = (m['animal'], m['phase'], m['date'])
            session = data_map.get(session_key)
            if not session:
                continue

            th, bi = PHASE_CONFIG[phase]
            cfg_s = dict(session)
            cfg_s['phase'] = phase
            cfg_s['threshold'] = th
            cfg_s['bidirectional'] = bi

            trials = _extract_trials(cfg_s)
            if not trials:
                continue

            outcomes = _extract_outcomes(trials)
            if not outcomes:
                continue

            pct = _get_pct(trials, outcomes)
            print(f'{animal}, {date_str} --> {pct:.2f}%')

            def _compute_window_success(out_slice):
                if not out_slice:
                    return None
                
                n_trials = len(out_slice)
                n_hits = sum(1 for o in out_slice if o == 'hit')

                return 100.0 * (n_hits / n_trials)

            N = 20
            w_num = 0
            win_x = []
            win_y = []

            start = 0
            while (start + N) < (len(outcomes) - 1):
                end = start + N
                w_outcomes = outcomes[start:end]
                s_val = _compute_window_success(w_outcomes)
                if s_val is not None:
                    w_num += 1
                    win_x.append(w_num)
                    win_y.append(s_val)
                start += 1

            if win_x and win_y:
                x_arr = np.asarray(win_x, float)
                y_arr = np.asarray(win_y, float)

                ax.plot(x_arr, y_arr,
                        linewidth=1.5,
                        markersize=5,
                        label=date_str
                        )
        
        ax.set_title(str(animal), fontsize=12, fontweight='bold')
        ax.set_xlabel('Window # (N = 20 trials)', fontsize=11, fontweight='normal')
        ax.set_ylabel('Success %', fontsize=11, fontweight='normal')
        ax.set_ylim(-10, 110)
        ax.set_yticks([0, 20, 40, 60, 80, 100])
        ax.legend(fontsize=6, loc='best')

        print('\n')
    
    for k in range(n_animals, len(axes_flat)):
        axes_flat[k].set_visible(False)

    fig.tight_layout()
    show_figure(fig)

    # plt.savefig(os.path.join(script_dir, 'windowed_success.png'), dpi=600)


# -----------------------------
# Main
# -----------------------------
if __name__ == '__main__':
    try:
        plot_success()
        # main()
        # run_pipeline()
    except KeyboardInterrupt:
        os.system('cls')
        print('\nAborted by user\n')
