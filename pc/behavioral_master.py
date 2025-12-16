import sys
import os
os.system('cls')
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = '1'

import warnings
warnings.filterwarnings('ignore', category=UserWarning, message='pkg_resources is deprecated as an API.*')

import time
from datetime import datetime
from collections import deque
from queue import Queue
import serial
import serial.tools.list_ports
from threading import Thread, Event, Lock
from openpyxl import load_workbook, Workbook
import json
import hashlib
import re
import requests
import shutil
import tempfile
import subprocess
from pathlib import Path
from cursor_utils import cursor_fcn

BAUDRATE = 1000000

SESSION_END_STRING = "S"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ANIMAL_MAP_PATH = os.path.join(SCRIPT_DIR, 'animal_map.json')

TMP_FILE = "session_data_tmp.json"
TMP_PATH = os.path.join(SCRIPT_DIR, TMP_FILE)
TMP_FLUSH_SEC = 150.0

SESSION_LOCK = Lock()
TMP_STOP = Event()
TMP_META = {}

EVT_QUEUE: "Queue[tuple[str, str]]" = Queue()
ENC_QUEUE: "Queue[tuple[str, str]]" = Queue()

REPO_URL = "https://github.com/GonzalesLabVU/Behavior-BCI.git"
REPO_OWNER = "GonzalesLabVU"
REPO_NAME = "Behavior-BCI"
REPO_BRANCH = "main"
REPO_XLSX_SUBDIR = "pc/data"
LOCK_REF_PREFIX = "refs/heads/write_flag_"
LOCK_POLL_SEC = 30
LOCK_MAX_WAIT_SEC = 600


# ----- STARTUP -----
def _load_animal_map(path=ANIMAL_MAP_PATH):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if not isinstance(data, dict):
            raise ValueError("'animal_map.json' file must contain an object mapping cohort -> list of animals")
    
        cohorts = {name: set(animals) for (name, animals) in data.items()}
    except FileNotFoundError:
        cohorts = {}
    except Exception as e:
        print(f'Warning: failed to load {path}: {e}. Starting with empty map')
        cohorts = {}
    
    return cohorts


def _save_animal_map(cohorts, path=ANIMAL_MAP_PATH):
    data = {name: sorted(list(animals)) for (name, animals) in cohorts.items()}

    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)


COHORTS = _load_animal_map()
ANIMAL_TO_COHORT = {
    animal: cohort_name
    for (cohort_name, animals) in COHORTS.items()
    for animal in animals
    }


# ----- GITHUB -----
def _gh_headers():
    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        raise RuntimeError("Missing GITHUB_TOKEN env var")
    
    return {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/vnd.github+json',
        'X-GitHub-Api-Version': '2022-11-28'
        }


def _gh_api(url, method='GET', json_body=None):
    r = requests.request(method, url, headers=_gh_headers(), json=json_body, timeout=20)
    return r


def _lock_ref_for(xlsx_basename):
    base = os.path.basename(xlsx_basename)
    stem = Path(base).stem

    slug = re.sub(r"[^A-Za-z0-9]+", "_", stem).strip("_")[:24]
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:12]

    return f'{LOCK_REF_PREFIX}{slug}_{digest}'


def _acquire_write_lock(lock_ref):
    base = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/git"
    url_create = f"{base}/refs"

    start = time.time()
    while True:
        r_main = _gh_api(f"{base}/ref/heads/{REPO_BRANCH}")
        if r_main.status_code != 200:
            raise RuntimeError(f'Failed to read branch ref: {r_main.status_code} {r_main.text}')
        
        main = r_main.json()
        sha = main['object']['sha']

        resp = _gh_api(url_create, method='POST', json_body={'ref': lock_ref, 'sha': sha})
        if resp.status_code in (201,):
            return
        
        if resp.status_code == 422:
            if time.time() - start > LOCK_MAX_WAIT_SEC:
                raise TimeoutError('Timed out waiting for write lock')
            
            time.sleep(LOCK_POLL_SEC)
            continue

        raise RuntimeError(f'Lock acquire failed: {resp.status_code} {resp.text}')


def _release_write_lock(lock_ref):
    base = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/git"
    url_delete = f"{base}/ref/{lock_ref.replace('refs/', '')}"

    resp = _gh_api(url_delete, method='DELETE')
    if resp.status_code in (204, 404):
        return
    
    raise RuntimeError(f'Lock release failed: {resp.status_code} {resp.text}')


def _git_run(args, cwd, *, capture=False):
    env = os.environ.copy()
    env["GIT_TERMINAL_PROMPT"] = "0"

    return subprocess.run(
        ["git", *args],
        cwd=str(cwd),
        check=True,
        text=True,
        capture_output=capture,
        env=env
        )


def _clone_repo_temp(parent_dir):
    tmp = Path(tempfile.mkdtemp(prefix="repo_tmp_", dir=parent_dir))

    try:
        _git_run(["clone", "--depth", "1", "--branch", REPO_BRANCH, REPO_URL, str(tmp)],
                 cwd=parent_dir)
    except Exception:
        shutil.rmtree(tmp, ignore_errors=True)
        raise

    return tmp


def _find_file_in_clone(clone_root, xlsx_basename):
    if REPO_XLSX_SUBDIR:
        candidate = clone_root / REPO_XLSX_SUBDIR / xlsx_basename
        if candidate.exists():
            return candidate
    
    matches = list(clone_root.rglob(xlsx_basename))
    matches = [p for p in matches if p.is_file()]

    if not matches:
        raise FileNotFoundError(f'{xlsx_basename} not found in repo clone')
    
    if len(matches) > 1:
        rels = [str(p.relative_to(clone_root)) for p in matches]
        raise RuntimeError(f'Multiple matches for {xlsx_basename} in repo clone: {rels}')
    
    return matches[0]


def _pull_from_repo(xlsx_basename):
    if not xlsx_basename.lower().endswith('.xlsx'):
        raise ValueError('Expected an .xlsx filename')
    
    clone_dir = _clone_repo_temp(SCRIPT_DIR)

    try:
        src = _find_file_in_clone(clone_dir, xlsx_basename)
        dst = Path(SCRIPT_DIR) / xlsx_basename
        shutil.copy2(src, dst)
        print(f'[git] Pulled {xlsx_basename} -> {dst}', flush=True)
    except FileNotFoundError:
        print(f'[git] {xlsx_basename} not found in repo, skipping pull', flush=True)
    finally:
        shutil.rmtree(clone_dir, ignore_errors=True)


def _push_to_repo(xlsx_basename):
    if not xlsx_basename.lower().endswith('.xlsx'):
        raise ValueError('Expected an .xlsx filename')
    
    local_file = Path(SCRIPT_DIR) / xlsx_basename
    if not local_file.exists():
        raise FileNotFoundError(f'Local file not found: {local_file}')
    
    token = os.environ.get('GITHUB_TOKEN')
    if not token:
        raise RuntimeError('Missing GITHUB_TOKEN env var')
    
    push_url = f'https://x-access-token:{token}@github.com/{REPO_OWNER}/{REPO_NAME}.git'
    clone_dir = _clone_repo_temp(SCRIPT_DIR)

    try:
        try:
            repo_target = _find_file_in_clone(clone_dir, xlsx_basename)
        except FileNotFoundError:
            if not REPO_XLSX_SUBDIR:
                raise

            repo_target = Path(clone_dir) / REPO_XLSX_SUBDIR / xlsx_basename
            repo_target.parent.mkdir(parents=True, exist_ok=True)

        shutil.copy2(local_file, repo_target)

        rel = repo_target.relative_to(clone_dir)

        _git_run(["add", str(rel)], cwd=clone_dir)

        st = _git_run(["status", "--porcelain", "--", str(rel)], cwd=clone_dir, capture=True)
        if not st.stdout.strip():
            print(f'[git] No changes to push for {xlsx_basename}', flush=True)
            return
        
        msg = f"Update {xlsx_basename} ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
        
        _git_run(["config", "user.email", "bci-bot@users.noreply.github.com"], cwd=clone_dir)
        _git_run(["config", "user.name", "BCI bot"], cwd=clone_dir)
        _git_run(["commit", "-m", msg], cwd=clone_dir)
        _git_run(["pull", "--rebase", "origin", REPO_BRANCH], cwd=clone_dir)
        _git_run(["push", push_url, f'HEAD:{REPO_BRANCH}'], cwd=clone_dir)

        print(f'[git] Committed {xlsx_basename} to {REPO_BRANCH}', flush=True)
    finally:
        shutil.rmtree(clone_dir, ignore_errors=True)


# ----- HELPERS -----
def _find_arduino_port():
    ports = serial.tools.list_ports.comports()
    for port in ports:
        dsc = (port.description or "").lower()

        if 'arduino' in dsc or 'usb serial' in dsc:
            return port.device
    return None


def _update_cohorts(animal_id):
    global COHORTS, ANIMAL_TO_COHORT

    print(f'\nAnimal {animal_id} is not currently assigned to a cohort.')
    choice = input('Would you like to assign it now? [y/N]:  ').strip().lower()

    if choice not in {'y', 'yes'}:
        return False
    
    while True:
        cohort_id = input('\nEnter cohort number for this animal (leave blank to cancel):  ')

        if not cohort_id:
            print('Cohort assignment cancelled\n')
            return False
        else:
            cohort_name = 'cohort' + cohort_id
        
        if cohort_name not in COHORTS:
            create = input(f'\nCohort {cohort_id} does not currently exist. Create it now? [y/N]:  ').strip().lower()

            if create not in {'y', 'yes'}:
                print('Cohort not created\n')
                continue

            COHORTS[cohort_name] = set()
        
        old_cohort = ANIMAL_TO_COHORT.get(animal_id)
        old_cohort_animals_before = set(COHORTS.get(old_cohort, set())) if old_cohort else None

        new_cohort_animals_before = set(COHORTS.get(cohort_name, set()))

        if old_cohort and old_cohort != cohort_name:
            COHORTS[old_cohort].discard(animal_id)
        
        COHORTS[cohort_name].add(animal_id)
        ANIMAL_TO_COHORT[animal_id] = cohort_name

        _save_animal_map(COHORTS)
        print(f'\nAssigned animal {animal_id} to cohort {cohort_id}\n')

        if old_cohort and old_cohort != cohort_name and old_cohort_animals_before is not None:
            _update_workbook_filename(old_cohort, old_cohort_animals_before, COHORTS.get(old_cohort, set()))

        _update_workbook_filename(cohort_name, new_cohort_animals_before, COHORTS.get(cohort_name, set()))
        
        return True


def _write_tmp(session_data):
    with SESSION_LOCK:
        evt = list(session_data.get('EVT', []))
        enc = list(session_data.get('ENC', []))
    
    payload = {
        'meta': dict(TMP_META),
        'event': {
            'timestamps': [x[0] for x in evt],
            'values': [x[1] for x in evt]
            },
        'encoder': {
            'timestamps': [x[0] for x in enc],
            'values': [x[1] for x in enc]
            }
        }
    
    tmp_path = TMP_PATH + ".tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=4)
    
    os.replace(tmp_path, TMP_PATH)


def _delete_tmp():
    try:
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)
    except Exception:
        pass


def _tmp_flush_loop(session_data):
    try:
        _write_tmp(session_data)
    except Exception:
        pass

    while not TMP_STOP.wait(TMP_FLUSH_SEC):
        try:
            _write_tmp(session_data)
        except Exception:
            pass


def _open_or_create_workbook(path, backup_path=None):
    folder = os.path.dirname(path)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
    
    def _load_or_new(p):
        if os.path.exists(p):
            return load_workbook(p)
        
        return Workbook()
    
    try:
        wb = _load_or_new(path)
        _validate_sheets(wb)
        
        return wb, False
    except Exception as e:
        if backup_path is None:
            wb = Workbook()
            _validate_sheets(wb)

            return wb, True
        
        print(f'\n[WARNING] Workbook load failed ({e}). Falling back to backup_data.xlsx\n', flush=True)

        try:
            wb = _load_or_new(backup_path)
        except Exception:
            wb = Workbook()
        
        _validate_sheets(wb)

        return wb, True


def _is_sheet_empty(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ''):
                return False
    
    return True


def _validate_sheets(wb):
    if 'Event' not in wb.sheetnames:
        wb.create_sheet('Event')
    if 'Encoder' not in wb.sheetnames:
        wb.create_sheet('Encoder')
    
    for name in list(wb.sheetnames):
        if name in ('Event', 'Encoder'):
            continue

        ws = wb[name]
        if _is_sheet_empty(ws):
            wb.remove(ws)


def _find_workbook_for_animal(animal_id):
    animal_id = animal_id.upper()

    cohort_name = ANIMAL_TO_COHORT.get(animal_id)
    if cohort_name is None:
        filename = 'backup_data.xlsx'
        print(f'\n[WARNING] Unknown animal {animal_id} or cohort {cohort_name}\n')
    else:
        animals_in_cohort = COHORTS.get(cohort_name, set())
        if animals_in_cohort:
            label = ''.join(sorted(animals_in_cohort))
        else:
            label = cohort_name
        
        filename = f'{label}_data.xlsx'
    
    return os.path.join(SCRIPT_DIR, filename)


def _update_workbook_filename(cohort_name, old_animals, new_animals):
    old_animals = set(old_animals or [])
    new_animals = set(new_animals or [])

    if not old_animals:
        return
    
    old_label = ''.join(sorted(old_animals))
    new_label = ''.join(sorted(new_animals)) if new_animals else cohort_name

    if old_label == new_label:
        return
    
    old_path = os.path.join(SCRIPT_DIR, f'{old_label}_data.xlsx')
    new_path = os.path.join(SCRIPT_DIR, f'{new_label}_data.xlsx')

    if not os.path.exists(old_path):
        return
    
    if os.path.exists(new_path):
        timestamp = int(time.time())
        alt_path = os.path.join(SCRIPT_DIR, f'{new_label}_data_{timestamp}.xlsx')

        new_base = os.path.basename(new_path)
        alt_base = os.path.basename(alt_path)

        os.rename(old_path, alt_path)
        print(f'\n[WARNING] {new_base} exists in the current folder; renamed old workbook to {alt_base}\n', flush=True)
    else:
        new_base = os.path.basename(new_path)
        old_base = os.path.basename(old_path)

        os.rename(old_path, new_path)
        print(f'\nRenamed workbook {old_base} to {new_base}\n', flush=True)


def _write_to_sheet(ws, date_str, animal_str, phase_str, data):
    col = 1
    while ws.cell(row=1, column=col).value is not None:
        col += 2
    
    ws.cell(row=1, column=col, value=date_str)
    ws.cell(row=2, column=col, value=animal_str)
    ws.cell(row=2, column=col+1, value=phase_str)

    for r, (ts, payload) in enumerate(data, start=4):
        try:
            val = float(payload)
        except Exception:
            val = str(payload)
        
        ws.cell(row=r, column=col, value=ts)
        ws.cell(row=r, column=col+1, value=val)


def _save_to_local(path, session_data, animal_id, phase_id, *, max_wait_sec=180, poll_sec=5):
    def _format_meta():
        session_date = datetime.now().date()
        date_str = f'{session_date.month}/{session_date.day}/{session_date.year}'
        animal_str = f'Animal {animal_id}' if animal_id is not None else 'Animal [unknown]'
        phase_str = f'Phase {phase_id}' if phase_id is not None else 'Phase [unknown]'

        return date_str, animal_str, phase_str
    
    target_path = path
    backup_path = os.path.join(SCRIPT_DIR, 'backup_data.xlsx')

    wb, backup_used = _open_or_create_workbook(target_path, backup_path=backup_path)
    if backup_used:
        target_path = backup_path
    
    date_str, animal_str, phase_str = _format_meta()

    try:
        with SESSION_LOCK:
            evt = list(session_data.get('EVT', []))
            enc = list(session_data.get('ENC', []))
    except Exception:
        evt = list(session_data.get('EVT', [])) if isinstance(session_data, dict) else []
        enc = list(session_data.get('ENC', [])) if isinstance(session_data, dict) else []
    
    try:
        if evt:
            ws = wb['Event'] if 'Event' in wb.sheetnames else wb.create_sheet('Event')
            _write_to_sheet(ws, date_str, animal_str, phase_str, evt)
        
        if enc:
            ws = wb['Encoder'] if 'Encoder' in wb.sheetnames else wb.create_sheet('Encoder')
            _write_to_sheet(ws, date_str, animal_str, phase_str, enc)
    except Exception as e:
        print(f"\n[ERROR] Failed while writing into workbook in memory: {e}\n", flush=True)
        return False, target_path, "write_failed"
    
    deadline = time.time() + max_wait_sec

    while True:
        try:
            wb.save(target_path)
            print(f"\nSaved data locally to: {target_path}\n", flush=True)
            return True, target_path, None

        except PermissionError:
            # Likely open in Excel
            remaining = int(max(0, deadline - time.time()))
            print(
                f"\n[WARNING] Cannot save because the workbook is locked (likely open in Excel).\n"
                f"Close Excel or disable protected view for this file, then save will retry.\n"
                f"Target: {target_path}\n"
                f"Raw backup retained: {TMP_PATH}\n"
                f"Retrying every {poll_sec}s for up to {remaining}s...\n",
                flush=True
            )
            if time.time() >= deadline:
                return False, target_path, "locked_by_excel"
            time.sleep(poll_sec)

        except Exception as e:
            print(f"\n[ERROR] wb.save failed for {target_path}: {e}\n"
                  f"Raw backup retained: {TMP_PATH}\n", flush=True)
            return False, target_path, "save_failed"


def _save_data(session_data, animal_id, phase_id):
    if animal_id not in ANIMAL_TO_COHORT:
        raise ValueError(f"Unknown animal ID: {animal_id}\n")

    save_path = _find_workbook_for_animal(animal_id)
    xlsx_base = os.path.basename(save_path)
    lock_ref = _lock_ref_for(xlsx_base)

    try:
        _acquire_write_lock(lock_ref)
    except TimeoutError:
        print(f"\n[WARNING] GitHub lock timed out for {xlsx_base}. Will save locally only.\n"
              f"Raw backup retained: {TMP_PATH}\n", flush=True)
        ok, saved_path, reason = _save_to_local(save_path, session_data, animal_id, phase_id)
        if ok:
            _delete_tmp()
        else:
            print(f"\n[WARNING] Local save issue ({reason}). Raw backup kept: {TMP_PATH}\n", flush=True)
        return
    except Exception as e:
        print(f"\n[WARNING] Lock acquisition failed for {xlsx_base} ({e}). Will save locally only.\n"
              f"Raw backup retained: {TMP_PATH}\n", flush=True)
        ok, saved_path, reason = _save_to_local(save_path, session_data, animal_id, phase_id)
        if ok:
            _delete_tmp()
        else:
            print(f"\n[WARNING] Local save issue ({reason}). Raw backup kept: {TMP_PATH}\n", flush=True)
        return

    try:
        _pull_from_repo(xlsx_base)

        ok, saved_path, reason = _save_to_local(save_path, session_data, animal_id, phase_id)
        if ok:
            _delete_tmp()
        else:
            print(f"\n[WARNING] Local save failed ({reason}). Not pushing.\n"
                  f"Raw backup retained: {TMP_PATH}\n", flush=True)
            return

        try:
            _push_to_repo(os.path.basename(saved_path))
            print(f"\n[git] Committed {os.path.basename(saved_path)} to {REPO_BRANCH}\n", flush=True)
        except Exception as e:
            print(f"\n[WARNING] Push failed ({e}).\n"
                  f"Data is saved locally to: {saved_path}\n"
                  f"Raw backup retained: {TMP_PATH}\n", flush=True)
    finally:
        try:
            _release_write_lock(lock_ref)
        except Exception as e:
            print(f"\n[WARNING] Lock release failed ({e}).\n", flush=True)


def _save_on_error(session_data, animal_id=None, phase_id=None):
    if not isinstance(session_data, dict):
        return

    try:
        with SESSION_LOCK:
            evt = list(session_data.get("EVT", []))
            enc = list(session_data.get("ENC", []))
    except Exception:
        evt = list(session_data.get("EVT", []))
        enc = list(session_data.get("ENC", []))

    if not evt and not enc:
        print("\nNo data collected\n", flush=True)
        return

    if animal_id is not None and animal_id in ANIMAL_TO_COHORT:
        save_path = _find_workbook_for_animal(animal_id)
    else:
        save_path = os.path.join(SCRIPT_DIR, "backup_data.xlsx")

    xlsx_base = os.path.basename(save_path)
    lock_ref = _lock_ref_for(xlsx_base)

    try:
        _acquire_write_lock(lock_ref)
        locked = True
    except TimeoutError:
        locked = False
        print(f"\n[WARNING] GitHub lock timed out during error-save for {xlsx_base}.\n"
              f"Will save locally only.\n"
              f"Raw backup retained: {TMP_PATH}\n", flush=True)
    except Exception as e:
        locked = False
        print(f"\n[WARNING] Lock acquisition failed during error-save for {xlsx_base} ({e}).\n"
              f"Will save locally only.\n"
              f"Raw backup retained: {TMP_PATH}\n", flush=True)

    try:
        if locked:
            _pull_from_repo(xlsx_base)

        ok, saved_path, reason = _save_to_local(save_path, session_data, animal_id, phase_id)
        if ok:
            _delete_tmp()
        else:
            print(f"\n[WARNING] Local save issue ({reason}). Raw backup kept: {TMP_PATH}\n", flush=True)
            return

        if locked:
            try:
                _push_to_repo(os.path.basename(saved_path))
                print(f"\n[git] Pushed {os.path.basename(saved_path)} successfully\n", flush=True)
            except Exception as e:
                print(f"\n[WARNING] Push failed ({e}).\n"
                      f"Data is saved locally to: {saved_path}\n"
                      f"Raw backup retained: {TMP_PATH}\n", flush=True)
    finally:
        if locked:
            try:
                _release_write_lock(lock_ref)
            except Exception as e:
                print(f"\n[WARNING] Lock release failed ({e}).\n", flush=True)


def _now_ts():
    return datetime.now().strftime("%H:%M:%S.%f")[:-3]


def _decode_line(raw):
    if not raw:
        return ""
    
    try:
        return raw.decode("utf-8", errors="strict").strip()
    except UnicodeDecodeError:
        return raw.decode("latin1", errors="ignore").strip()


def _handle_line(line, session_data, N=None, trial_stack=None):
    if not line:
        return None
    
    if line == SESSION_END_STRING:
        return 'S'
    
    if line.startswith("[EVT]"):
        parts = line.split("]", 1)
        payload = parts[1].strip() if len(parts) > 1 else ""
        ts = _now_ts()

        with SESSION_LOCK:
            session_data["EVT"].append([ts, payload])

        EVT_QUEUE.put((ts, payload))

        if payload in {"hit", "miss"}:
            with SESSION_LOCK:
                if len(session_data['EVT']) >= 2:
                    prev_payload = session_data['EVT'][-2][1]
                    if prev_payload in {"hit", "miss"}:
                        return None
            
            if trial_stack is not None:
                if len(trial_stack) >= N:
                    trial_stack.pop()
                trial_stack.appendleft(payload)

        return None
    
    if line.startswith("[ENC]"):
        parts = line.split("]", 1)
        payload = parts[1].strip() if len(parts) > 1 else ""
        ts = _now_ts()
        
        with SESSION_LOCK:
            session_data["ENC"].append([ts, payload])

        ENC_QUEUE.put((ts, payload))

        return None

    return None


def _wait_for_ack(ser, session_data, timeout=2.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if ser.in_waiting > 0:
            raw = ser.readline()
            line = _decode_line(raw)
            if not line:
                continue

            if line == 'A':
                return True
            
            _handle_line(line, session_data)
        else:
            time.sleep(0.005)
    
    return False


def _send_to_serial(ser, session_data, text, timeout=2.0):
    ser.write((text.strip() + "\n").encode("utf-8"))
    ser.flush()

    ack_received = _wait_for_ack(ser, session_data, timeout=timeout)
    if not ack_received:
        raise TimeoutError(f'No ACK after sending: {text!r}')


def _set_easy_rate(ser, session_data, trial_stack):
    n_hits = sum(1 for x in trial_stack if x == 'hit')
    if n_hits < 10:
        K = 3
    elif n_hits == 10:
        K = 5
    else:
        K = 7
    N = 4 * K

    trial_stack.clear()

    try:
        _send_to_serial(ser, session_data, str(K))
    except Exception as e:
        print(f"Failed to send calibrated easy-trial rate to Arduino: {e}\n")
        
        try:
            ser.close()
        except Exception:
            pass

        raise RuntimeError
    
    EVT_QUEUE.put((_now_ts(), f'setK {K}'))

    return K, N, n_hits


def _is_early_exit(trial_stack, N):
    if len(trial_stack) < N:
        return False

    n_hits = sum(1 for x in trial_stack if x == 'hit')
    return n_hits < 7


def _print_window(trial_stack, N, trial_index):
    if N is None or N <= 0:
        return
    
    if len(trial_stack) < N:
        return
    
    n_hits = sum(1 for x in trial_stack if x == 'hit')
    hit_rate = 100.0 * (n_hits / N)

    end_trial = trial_index
    start_trial = (end_trial - N + 1)

    print(f'(Trials {start_trial}-{end_trial}):  {n_hits}/{N} ({hit_rate:.1f}%)', flush=True)


def _terminate_session(ser, session_data, msg=None):
    try:
        ser.write(("E\n").encode("utf-8"))
        ser.flush()

        end_deadline = time.time() + 2.0
        while time.time() < end_deadline:
            if ser.in_waiting > 0:
                raw = ser.readline()
                line = _decode_line(raw)
                flag = _handle_line(line, session_data)

                if flag == 'S':
                    print(f"\n\n{msg or 'Session terminated'}\n")
                    break
            else:
                time.sleep(0.01)
    except Exception as e:
        print(f'\n\nFailed to terminate session: {e}\n')
    finally:
        if ser and ser.is_open:
            ser.close()


# -------- TOP LEVEL --------
def setup(session_data):
    # initialize serial connection with Arduino
    port = _find_arduino_port()
    if not port:
        print("No Arduino detected\n", flush=True)
        sys.exit(1)
    
    try:
        ser = serial.Serial(port, BAUDRATE, timeout=0.05)
        time.sleep(2)
        print(f"\nConnected to {port}\n", flush=True)
    except serial.SerialException as e:
        print(f"\nSerial error: {e}\n")
        sys.exit(1)
    
    # ask for animal ID
    dev_mode = False

    while True:
        try:
            animal_id_raw = input("\nAnimal ID: ").strip().upper()

            if animal_id_raw == "":
                animal_id = 'DEV'
                dev_mode = True
                print("\n------------\n  DEV MODE \n------------\n")
                break
            else:
                animal_id = animal_id_raw

            if animal_id and animal_id in ANIMAL_TO_COHORT:
                break

            if animal_id:
                if _update_cohorts(animal_id):
                    break

            print("Please enter a valid ID\n")
        except KeyboardInterrupt:
            print('\nProcess terminated by user\n')
            sys.exit(0)
    
    # ask for phase ID
    while True:
        try:
            phase_id = input("Training Phase: ").strip()

            if phase_id and phase_id in {"1", "2", "3", "4", "5"}:
                break
            print("Please enter a valid training phase\n")
        except KeyboardInterrupt:
            print('\nProcess terminated by user\n')
            sys.exit(0)
    
    # send phase ID to Arduino
    try:
        _send_to_serial(ser, session_data, phase_id)
    except Exception as e:
        print(f"Failed to send training phase to Arduino: {e}\n")
        ser.close()
        sys.exit(1)
    
    # start cursor thread if needed
    cursor_thread = None

    if str(phase_id) not in {'1', '2'}:
        threshold = {
            '3': 30.0,
            '4': 60.0,
            '5': 90.0
            }[str(phase_id)]
        cursor_thread = Thread(
            target=cursor_fcn,
            args=(threshold, EVT_QUEUE, ENC_QUEUE),
            kwargs={'display_index': 1, 'fullscreen': True},
            daemon=True
            )
        
        cursor_thread.start()
    
    return ser, animal_id, phase_id, session_data, cursor_thread, dev_mode


def main(ser, session_data, phase_id):
    K = 5
    N = 20

    try:
        _send_to_serial(ser, session_data, str(K))
    except Exception as e:
        print(f"Failed to send initial easy-trial rate to Arduino: {e}\n")
        ser.close()
        sys.exit(1)
    
    EVT_QUEUE.put((_now_ts(), f'setK {K}'))

    trial_stack = deque()
    calibrated = False
    finished = False
    t_start = None
    t_stop = None
    trial_idx = 0

    try:
        while ser and ser.is_open and not finished:
            while ser.in_waiting > 0:
                if t_start is None:
                    t_start = time.time()

                    print(f"\nSession start time: {datetime.now().strftime('%I:%M %p')}")
                    print("\nRunning session protocol...", end="", flush=True)

                raw = ser.readline()
                line = _decode_line(raw)
                flag = _handle_line(line, session_data, N, trial_stack)

                if phase_id not in {'1', '2'} and calibrated:
                    trial_idx += 1
                    _print_window(trial_stack, N, trial_idx)

                if phase_id not in {'1', '2'}:
                    if not calibrated:
                        if len(trial_stack) >= N:
                            K, N, calibration_hits = _set_easy_rate(ser, session_data, trial_stack)
                            calibrated = True

                            print(f'\nCalibration period finished [hits={calibration_hits}/20, K={K}, N={N}]\n', flush=True)
                    else:
                        if _is_early_exit(trial_stack, N):
                            _terminate_session(ser, session_data, msg='Session terminated by early exit')

                            t_stop = time.time()
                            finished = True
                            break
                
                if flag == 'S':
                    t_stop = time.time()
                    finished = True
                    break
            
            time.sleep(0.01)

    except KeyboardInterrupt:
        _terminate_session(ser, session_data, msg='Session terminated by user')

        t_stop = time.time()
        finished = True

    finally:
        if ser and ser.is_open:
            ser.close()
    
    if t_stop is None:
        t_stop = time.time()
    
    if t_start is None:
        t_start = t_stop

    elapsed = int(t_stop - t_start)
    m, s = divmod(elapsed, 60)
    print(f'\nSession duration: {m:02d}:{s:02d}')

    return session_data


if __name__ == "__main__":
    session_data = {"EVT": deque(), "ENC": deque()}
    animal_id = None
    phase_id = None
    dev_mode = True
    ser = None
    cursor_thread = None
    tmp_thread = None

    try:
        ser, animal_id, phase_id, session_data, cursor_thread, dev_mode = setup(session_data)

        TMP_META = {
            'date': datetime.now().strftime('%m/%d/%Y'),
            'animal': animal_id,
            'phase': phase_id
            }
        
        tmp_thread = Thread(target=_tmp_flush_loop, args=(session_data,), daemon=True)
        tmp_thread.start()

        session_data = main(ser, session_data, phase_id)

        try:
            import pygame

            if pygame.get_init():
                pygame.event.post(pygame.event.Event(pygame.QUIT))
        except Exception:
            pass

        if cursor_thread is not None:
            cursor_thread.join(timeout=2.0)
        
        if not dev_mode and animal_id in ANIMAL_TO_COHORT:
            _save_data(session_data, animal_id, phase_id)
            
        # [plotting logic]

    except KeyboardInterrupt:
        _save_on_error(session_data, animal_id, phase_id)

    except SystemExit:
        try:
            with SESSION_LOCK:
                has_any = bool(session_data.get('EVT')) or bool (session_data.get('ENC'))
            
            if has_any:
                _save_on_error(session_data, animal_id, phase_id)
        finally:
            raise

    except Exception as e:
        _save_on_error(session_data, animal_id, phase_id)
    
    finally:
        TMP_STOP.set()

        if tmp_thread is not None:
            tmp_thread.join(timeout=2.0)
